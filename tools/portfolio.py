#!/usr/bin/env python
"""MYRA Portfolio Tracker - CLI tool for managing stock holdings.

WARNING: This tool accesses your financial holdings.
Never commit myra_portfolio.db or exports/ to git.
"""

import sys
import os
import csv
import json
import re
import sqlite3
from datetime import datetime, timedelta

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    from myra_app.portfolio_db import (
        get_db_path,
        get_connection as get_portfolio_conn,
        import_holdings,
        get_all_holdings,
        get_holding,
        add_holding,
        update_holding,
        delete_holding,
        record_snapshot,
        get_snapshots,
        get_transactions,
        get_delivery_metrics,
        get_technical_position,
        get_sector_allocation,
        get_scanner_overlap,
        get_delivery_alerts,
        get_concentration_risk,
        get_drawdown_metrics,
        get_allocation_by_mcap,
        get_volatility_metrics,
        get_diversification_score,
        _get_portfolio_meta,
        auto_refresh_portfolio,
    )
    from myra_app.constants import DB_DIR
except ImportError as e:
    print(f"Error: cannot import MYRA modules: {e}")
    sys.exit(1)

try:
    from tabulate import tabulate
except ImportError:
    tabulate = None

try:
    import yfinance as yf
except ImportError:
    yf = None

import pandas as pd


GREEN = "\033[92m"
RED = "\033[91m"
YELLOW = "\033[93m"
CYAN = "\033[96m"
BOLD = "\033[1m"
RESET = "\033[0m"

try:
    "₹".encode(sys.stdout.encoding or "utf-8")
    CURRENCY = "₹"
except (UnicodeEncodeError, UnicodeDecodeError, AttributeError):
    CURRENCY = "Rs."

_NEEDS_ASCII = sys.platform == "win32" or (sys.stdout.encoding or "").lower() in (
    "cp1252",
    "ascii",
)

if _NEEDS_ASCII:
    ARROW_UP = "^"
    ARROW_DOWN = "v"
    ARROW_RIGHT = "->"
    DASH = "-"
    CHECK = "Y"
    TRIANGLE_UP = "^"
    TRIANGLE_DOWN = "v"
    DIAMOND = "<>"
else:
    ARROW_UP = "\u2191"
    ARROW_DOWN = "\u2193"
    ARROW_RIGHT = "\u2192"
    DASH = "\u2014"
    CHECK = "\u2713"
    TRIANGLE_UP = "\u25b2"
    TRIANGLE_DOWN = "\u25bc"
    DIAMOND = "\u25c6"

EXPORTS_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "exports"
)


def fmt_inr(amount):
    if amount is None:
        return f"{YELLOW}N/A{RESET}"
    sign = ""
    s = f"{abs(amount):.2f}"
    int_part, _, dec_part = s.partition(".")
    last_three = int_part[-3:]
    rest = int_part[:-3]
    if rest:
        groups = []
        while len(rest) > 2:
            groups.append(rest[-2:])  # noqa: PG-APPEND
            rest = rest[:-2]
        if rest:
            groups.append(rest)
        groups.reverse()
        result = ",".join(groups) + "," + last_three
    else:
        result = last_three
    prefix = "-" if amount < 0 else ""
    return f"{prefix}{CURRENCY}{result}.{dec_part}"


def color_pnl(val, show_sign=True):
    if val is None:
        return f"{YELLOW}N/A{RESET}"
    if val >= 0:
        return f"{GREEN}{fmt_inr(val) if show_sign else ''}{RESET}"
    return f"{RED}{fmt_inr(val) if show_sign else ''}{RESET}"


def color_pnl_pct(val):
    if val is None:
        return f"{YELLOW}N/A{RESET}"
    if val >= 0:
        return f"{GREEN}+{val:.2f}%{RESET}"
    return f"{RED}{val:.2f}%{RESET}"


def color_price(val, high_good=False):
    if val is None:
        return f"{YELLOW}N/A{RESET}"
    if high_good:
        if val >= 0:
            return f"{GREEN}+{val:.1f}%{RESET}"
        return f"{RED}{val:.1f}%{RESET}"
    if val >= 0:
        return f"{GREEN}+{val:.1f}%{RESET}"
    return f"{RED}{val:.1f}%{RESET}"


def styled(text, style):
    return f"{style}{text}{RESET}"


def get_last_close(symbol):
    db_path = os.path.join(DB_DIR, "myra_technical.db")
    if not os.path.exists(db_path):
        return None
    try:
        conn = sqlite3.connect(db_path)
        cur = conn.execute(
            "SELECT close, date FROM technical_data WHERE symbol=? ORDER BY date DESC LIMIT 1",
            (symbol,),
        )
        row = cur.fetchone()
        conn.close()
        if row:
            return {"price": row[0], "date": row[1]}
        return None
    except sqlite3.Error:
        return None


def get_live_price(symbol):
    if yf is None:
        return None, "yfinance not installed"
    try:
        ticker = yf.Ticker(f"{symbol}.NS")
        hist = ticker.history(period="1d", interval="1m")
        if hist.empty:
            hist = ticker.history(period="5d")
        if hist.empty:
            return None, "No data from yfinance"
        return hist["Close"].iloc[-1], None
    except Exception as e:
        return None, str(e)


def get_fundamentals(symbol):
    db_path = os.path.join(DB_DIR, "myra_valuation.db")
    if not os.path.exists(db_path):
        return {}
    try:
        conn = sqlite3.connect(db_path)
        cur = conn.execute(
            "SELECT pe, sector, market_cap, industry, roe FROM fundamentals WHERE symbol=?",
            (symbol,),
        )
        row = cur.fetchone()
        conn.close()
        if row:
            return {
                "pe": row[0],
                "sector": row[1],
                "market_cap": row[2],
                "industry": row[3],
                "roe": row[4],
            }
        return {}
    except sqlite3.Error:
        return {}


def fmt_del_trend(trend):
    if trend == ARROW_UP:
        return f"{GREEN}{trend}{RESET}"
    if trend == ARROW_DOWN:
        return f"{RED}{trend}{RESET}"
    return trend


def fmt_vs_sma(pct):
    if pct is None:
        return f"{YELLOW}{DASH}{RESET}"
    if pct >= 0:
        return f"{GREEN}+{pct}% ABOVE{RESET}"
    return f"{RED}{pct}% BELOW{RESET}"


def fmt_vs_52w_high(pct):
    if pct is None:
        return f"{YELLOW}{DASH}{RESET}"
    return f"{styled(f'{pct}% FROM HI', RED if pct < -20 else YELLOW)}"


def fmt_vs_52w_low(pct):
    if pct is None:
        return f"{YELLOW}{DASH}{RESET}"
    return f"{styled(f'+{pct}% FROM LO', GREEN if pct > 20 else YELLOW)}"


def fmt_alert_short(alerts_list):
    for a in alerts_list:
        t = a["alert_type"]
        if t == "DELIVERY SURGE":
            return f"{GREEN}{TRIANGLE_UP} SURGE{RESET}"
        if t == "DELIVERY COLLAPSE":
            return f"{RED}!! COLLAPSE{RESET}"
        if t == "ABSORPTION":
            return f"{CYAN}{DIAMOND} ABSORB{RESET}"
        if t == "DISTRIBUTION":
            return f"{RED}{TRIANGLE_DOWN} DISTRIB{RESET}"
    return f"{YELLOW}{DASH}{RESET}"


def match_columns(df, candidates):
    for col in candidates:
        if col in df.columns:
            return col
    lower = {c.lower().replace(" ", "_").replace(".", ""): c for c in df.columns}
    for col in candidates:
        key = col.lower().replace(" ", "_").replace(".", "")
        if key in lower:
            return lower[key]
    return None


def build_portfolio_rows(use_live, sort_col):
    holdings = get_all_holdings()
    if not holdings:
        print(f"{YELLOW}No holdings in portfolio. Use 'import' or 'add' first.{RESET}")
        sys.exit(0)
    all_alerts = get_delivery_alerts(holdings)
    alert_map = {}
    for a in all_alerts:
        alert_map.setdefault(a["symbol"], []).append(a)  # noqa: PG-APPEND
    scanner_map = get_scanner_overlap(holdings)
    rows = []
    for h in holdings:
        eod = get_last_close(h["symbol"])
        price = eod["price"] if eod else None
        price_date = eod["date"] if eod else None
        if use_live:
            live_price, err = get_live_price(h["symbol"])
            if live_price is not None:
                price = live_price
            else:
                print(
                    f"{YELLOW}Live price failed for {h['symbol']}: {err} - using EOD{RESET}"
                )
        invested = h["net_qty"] * h["avg_price"]
        current_value = h["net_qty"] * price if price else 0
        overall_pnl = current_value - invested
        overall_pnl_pct = (overall_pnl / invested * 100) if invested else 0
        day_pnl = 0
        day_pnl_pct = 0.0
        if price and eod and price_date:
            conn = sqlite3.connect(os.path.join(DB_DIR, "myra_technical.db"))
            try:
                prev = conn.execute(  # noqa: PG-NPLUS1
                    "SELECT close FROM technical_data WHERE symbol=? AND date < ? ORDER BY date DESC LIMIT 1",
                    (h["symbol"], price_date),
                ).fetchone()
                if prev:
                    day_pnl = h["net_qty"] * (price - prev[0])
                    day_pnl_pct = (
                        (day_pnl / current_value * 100) if current_value else 0
                    )
            except sqlite3.Error:
                pass
            conn.close()
        fund = get_fundamentals(h["symbol"])
        del_m = get_delivery_metrics(h["symbol"])
        tech_pos = get_technical_position(h["symbol"])
        row = {
            "symbol": h["symbol"],
            "qty": h["net_qty"],
            "avg": h["avg_price"],
            "ltp": price or 0,
            "invested": invested,
            "current": current_value,
            "pnl": overall_pnl,
            "pnl_pct": overall_pnl_pct,
            "day_pnl": day_pnl,
            "day_pnl_pct": day_pnl_pct,
            "pe": fund.get("pe"),
            "sector": fund.get("sector", ""),
            "market_cap": fund.get("market_cap"),
            "roe": fund.get("roe"),
            "del_pct": del_m["del_pct"] if del_m else None,
            "del_trend": del_m["del_trend"] if del_m else None,
            "avg_del_20d": del_m["avg_del_20d"] if del_m else None,
            "vs_sma_pct": tech_pos["vs_sma_pct"] if tech_pos else None,
            "vs_52w_high_pct": tech_pos["vs_52w_high_pct"] if tech_pos else None,
            "vs_52w_low_pct": tech_pos["vs_52w_low_pct"] if tech_pos else None,
            "alerts": alert_map.get(h["symbol"], []),
            "scanner": scanner_map.get(h["symbol"], {}),
        }
        if use_live:
            row["live"] = live_price
        rows.append(row)  # noqa: PG-APPEND
    if sort_col:
        key_map = {
            "symbol": lambda r: r["symbol"],
            "value": lambda r: r["current"],
            "pnl": lambda r: r["pnl"],
            "day_pnl": lambda r: r["day_pnl"],
            "pnl_pct": lambda r: r["pnl_pct"],
        }
        key_fn = key_map.get(sort_col)
        if key_fn:
            rows.sort(key=key_fn, reverse=(sort_col != "symbol"))
    return rows


def _parse_indian_num(val):
    if val is None:
        return None
    s = str(val).strip().replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def _parse_pnl(val):
    if val is None:
        return None
    s = str(val).strip().replace(",", "")
    m = re.match(r"([+-]?[\d.]+)", s)
    if m:
        return float(m.group(1))
    return None


COL_MAP = {
    "symbol": ["Symbol"],
    "net_qty": ["Net Qty", "NetQty", "Quantity", "Qty", "BQTY"],
    "avg_price": [
        "Avg. Price",
        "Avg Price",
        "Average Price",
        "AvgPrice",
        "Buy Price",
        "Price",
    ],
    "ltp": ["LTP", "Last Price", "Close"],
    "category": ["Category", "Exchange", "Segment"],
}


def _find_header_row(sheet_data):
    for idx, row in enumerate(sheet_data):
        if row and str(row[0]).strip().startswith("Symbol"):
            return idx
    return None


def _build_col_map(header_row):
    col_index = {}
    for col_idx, cell in enumerate(header_row):
        raw = str(cell).strip() if cell else ""
        for internal, candidates in COL_MAP.items():
            for c in candidates:
                if raw == c or raw.startswith(c + " ("):
                    col_index[internal] = col_idx
                    break
    return col_index


def cmd_import(args):
    path = args.path
    if not os.path.exists(path):
        print(f"{RED}File not found:{RESET} {path}")
        sys.exit(1)

    try:
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        print(f"{RED}Failed to read XLSX:{RESET} {e}")
        sys.exit(1)

    sheet_name = None
    for prefix in ("Demat", "All"):
        for candidate in wb.sheetnames:
            if candidate.startswith(prefix):
                sheet_name = candidate
                break
        if sheet_name:
            break
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    sheet_data = list(ws.iter_rows(values_only=True))
    wb.close()

    header_row_idx = _find_header_row(sheet_data)
    if header_row_idx is None:
        print(f"{RED}Could not find header row in sheet '{sheet_name}'.{RESET}")
        print(
            f"  First cells: {[str(r[0])[:30] for r in sheet_data[:6] if r and r[0]]}"
        )
        sys.exit(1)

    header = list(sheet_data[header_row_idx])
    col_map = _build_col_map(header)
    if "symbol" not in col_map:
        print(f"{RED}Could not identify Symbol column in sheet '{sheet_name}'.{RESET}")
        print(f"  Header row: {header}")
        sys.exit(1)

    data_rows = []
    for row in sheet_data[header_row_idx + 1 :]:
        if not row or not row[col_map["symbol"]]:
            continue
        symbol = str(row[col_map["symbol"]]).strip().upper()
        if not symbol:
            continue
        qty_raw = (
            str(row[col_map.get("net_qty", 0)]).strip()
            if col_map.get("net_qty") is not None
            else "0"
        )
        price_raw = (
            str(row[col_map.get("avg_price", 0)]).strip()
            if col_map.get("avg_price") is not None
            else "0"
        )
        qty_val = _parse_indian_num(qty_raw)
        price_val = _parse_indian_num(price_raw)
        if qty_val is None or price_val is None or qty_val <= 0 or price_val <= 0:
            continue
        category = "NSE EQ"
        if col_map.get("category") is not None:
            cat_val = row[col_map["category"]]
            if cat_val:
                category = str(cat_val).strip()
        data_rows.append(  # noqa: PG-APPEND
            {
                "symbol": symbol,
                "net_qty": int(qty_val),
                "avg_price": price_val,
                "category": category,
            }
        )

    if not data_rows:
        print(f"{RED}No valid holdings found in sheet '{sheet_name}'.{RESET}")
        sys.exit(1)

    summary = {}
    if header_row_idx >= 2 and sheet_data[0] and sheet_data[1]:
        summary["invested"] = _parse_indian_num(sheet_data[1][0])  # noqa: PG-CHAINED
        summary["current"] = _parse_indian_num(sheet_data[1][1])  # noqa: PG-CHAINED
        summary["overall_pnl"] = _parse_pnl(sheet_data[1][2])  # noqa: PG-CHAINED
        summary["day_pnl"] = _parse_pnl(sheet_data[1][3])  # noqa: PG-CHAINED

    count = import_holdings(data_rows)
    print(f"{GREEN}Imported {count} holdings from {sheet_name}{RESET}")
    print(f"{DASH * 53}")
    if summary.get("invested") is not None:
        overall_pnl = summary.get("overall_pnl")
        overall_pnl_str = (
            f" ({color_pnl_pct(overall_pnl / summary['invested'] * 100) if overall_pnl is not None and summary['invested'] else ''})"
            if overall_pnl is not None
            else ""
        )
        print(f"Invested:    {fmt_inr(summary['invested'])}")
        print(f"Current:     {fmt_inr(summary['current'])}")
        print(
            f"Overall P&L: {color_pnl(summary['overall_pnl']) if summary['overall_pnl'] is not None else 'N/A'}{overall_pnl_str}"
        )
        print(
            f"Day P&L:     {color_pnl(summary['day_pnl']) if summary['day_pnl'] is not None else 'N/A'}"
        )
    else:
        invested = sum(r["net_qty"] * r["avg_price"] for r in data_rows)
        print(f"Invested:    {fmt_inr(invested)}")
    print(f"{DASH * 53}")


def cmd_view(args):
    if args.live:
        print(
            f"{YELLOW}Live prices via yfinance. Data may be delayed by 15 minutes. Use for reference only.{RESET}"
        )
    rows = build_portfolio_rows(args.live, args.sort)
    if not rows:
        return
    total_invested = sum(r["invested"] for r in rows)
    total_current = sum(r["current"] for r in rows)
    total_pnl = total_current - total_invested
    total_pnl_pct = (total_pnl / total_invested * 100) if total_invested else 0
    total_day_pnl = sum(r["day_pnl"] for r in rows)
    print(f"\n{BOLD}Portfolio Summary{RESET}")
    print(f"  Total Invested: {fmt_inr(total_invested)}")
    print(f"  Total Current:  {fmt_inr(total_current)}")
    print(f"  Overall P&L:    {color_pnl(total_pnl)} ({color_pnl_pct(total_pnl_pct)})")
    print(f"  Day P&L:        {color_pnl(total_day_pnl)}")
    if args.compact:
        headers = ["Symbol", "LTP", "P&L%", "Alert"]
        table_data = []
        for r in rows:
            table_data.append(  # noqa: PG-APPEND
                [
                    r["symbol"],
                    fmt_inr(r["ltp"]),
                    color_pnl_pct(r["pnl_pct"]),
                    fmt_alert_short(r["alerts"]),
                ]
            )

    def _fmt_del_trend(trend):
        if not trend or trend == DASH:
            return f"{YELLOW}{DASH}{RESET}"
        t = trend.strip()
        if t in ("\u2191", ARROW_UP):
            return f"{GREEN}{ARROW_UP}{RESET}"
        if t in ("\u2193", ARROW_DOWN):
            return f"{RED}{ARROW_DOWN}{RESET}"
        return f"{YELLOW}{ARROW_RIGHT}{RESET}"

    def _fmt_del_pct(val):
        if val is None:
            return f"{YELLOW}{DASH}{RESET}"
        return f"{val}%"

    if args.compact:
        headers = ["Symbol", "LTP", "P&L%", "Alert"]
        table_data = []
        for r in rows:
            table_data.append(  # noqa: PG-APPEND
                [
                    r["symbol"],
                    fmt_inr(r["ltp"]),
                    color_pnl_pct(r["pnl_pct"]),
                    fmt_alert_short(r["alerts"]),
                ]
            )
    elif args.detailed:
        headers = [
            "Symbol",
            "LTP",
            "P&L%",
            "Del%",
            "Trend",
            "vs 50SMA",
            "vs 52wH",
            "vs 52wL",
            "Alert",
        ]
        table_data = []
        for r in rows:
            table_data.append(  # noqa: PG-APPEND
                [
                    r["symbol"],
                    fmt_inr(r["ltp"]),
                    color_pnl_pct(r["pnl_pct"]),
                    _fmt_del_pct(r["del_pct"]),
                    _fmt_del_trend(r["del_trend"]),
                    fmt_vs_sma(r["vs_sma_pct"]),
                    fmt_vs_52w_high(r["vs_52w_high_pct"]),
                    fmt_vs_52w_low(r["vs_52w_low_pct"]),
                    fmt_alert_short(r["alerts"]),
                ]
            )
    else:
        headers = [
            "Symbol",
            "Qty",
            "Avg",
            "LTP",
            "Invested",
            "Current",
            "P&L",
            "P&L%",
            "Del%",
            "Trend",
            "Alert",
        ]
        table_data = []
        for r in rows:
            table_data.append(  # noqa: PG-APPEND
                [
                    r["symbol"],
                    r["qty"],
                    fmt_inr(r["avg"]),
                    fmt_inr(r["ltp"]),
                    fmt_inr(r["invested"]),
                    fmt_inr(r["current"]),
                    color_pnl(r["pnl"]),
                    color_pnl_pct(r["pnl_pct"]),
                    _fmt_del_pct(r["del_pct"]),
                    _fmt_del_trend(r["del_trend"]),
                    fmt_alert_short(r["alerts"]),
                ]
            )
    if tabulate:
        print(f"\n{tabulate(table_data, headers=headers, tablefmt='simple')}")
    else:
        print(f"\n{'  '.join(headers)}")
        for row in table_data:
            print("  ".join(str(c) for c in row))


def cmd_add(args):
    symbol = args.symbol.upper()
    qty = args.qty
    price = args.price
    category = args.category
    existing = get_holding(symbol)
    if existing:
        print(
            f"{YELLOW}{symbol} already exists ({existing['net_qty']} shares @ {fmt_inr(existing['avg_price'])}).{RESET}"
        )
        confirm = input("Overwrite? (y/N): ").strip().lower()
        if confirm != "y":
            print("Aborted.")
            return
    add_holding(symbol, qty, price, category)
    print(f"{GREEN}Added {symbol}: {qty} shares @ {fmt_inr(price)}{RESET}")
    if existing:
        print("Transaction logged as BUY (overwrite).")


def cmd_sell(args):
    symbol = args.symbol.upper()
    qty = args.qty
    price = args.price
    holding = get_holding(symbol)
    if not holding:
        print(f"{RED}Holding not found: {symbol}{RESET}")
        return
    if qty > holding["net_qty"]:
        print(f"{RED}Cannot sell {qty} shares - only {holding['net_qty']} held.{RESET}")
        return
    realised_pnl = qty * (price - holding["avg_price"])
    if qty == holding["net_qty"]:
        confirm = input(f"Sell ALL {qty} shares of {symbol}? (y/N): ").strip().lower()
        if confirm != "y":
            print("Aborted.")
            return
        delete_holding(symbol)
    else:
        update_holding(symbol, net_qty=holding["net_qty"] - qty)
        conn = get_portfolio_conn()
        conn.execute(
            """
            INSERT INTO transactions (symbol, action, qty, price, notes)
            VALUES (?, 'SELL', ?, ?, ?)
        """,
            (symbol, qty, price, f"Sold {qty} @ {price}"),
        )
        conn.commit()
        conn.close()
    pnl_str = color_pnl(realised_pnl)
    print(f"{GREEN}Sold {qty} shares of {symbol} @ {fmt_inr(price)}{RESET}")
    print(f"  Realised P&L: {pnl_str}")


def cmd_update(args):
    symbol = args.symbol.upper()
    holding = get_holding(symbol)
    if not holding:
        print(f"{RED}Holding not found: {symbol}{RESET}")
        return
    updates = {}
    if args.ltp is not None:
        updates["avg_price"] = args.ltp
    if args.qty is not None:
        updates["net_qty"] = args.qty
    if args.avg is not None:
        updates["avg_price"] = args.avg
    if not updates:
        print(f"{YELLOW}Nothing to update. Use --ltp, --qty, or --avg.{RESET}")
        return
    print(f"Updating {symbol}:")
    for k, v in updates.items():
        print(f"  {k}: {holding.get(k, 'N/A')} -> {v}")
    confirm = input("Proceed? (y/N): ").strip().lower()
    if confirm != "y":
        print("Aborted.")
        return
    update_holding(symbol, **updates)
    print(f"{GREEN}{symbol} updated.{RESET}")


def cmd_snapshot(args):
    holdings = get_all_holdings()
    if not holdings:
        print(f"{YELLOW}No holdings - cannot take snapshot.{RESET}")
        return
    total_invested = 0
    total_current = 0
    day_pnl = 0
    for h in holdings:
        eod = get_last_close(h["symbol"])
        price = eod["price"] if eod else 0
        invested = h["net_qty"] * h["avg_price"]
        current = h["net_qty"] * price
        total_invested += invested
        total_current += current
        if eod and eod["date"]:
            conn = sqlite3.connect(os.path.join(DB_DIR, "myra_technical.db"))
            try:
                prev = conn.execute(  # noqa: PG-NPLUS1
                    "SELECT close FROM technical_data WHERE symbol=? AND date < ? ORDER BY date DESC LIMIT 1",
                    (h["symbol"], eod["date"]),
                ).fetchone()
                if prev:
                    day_pnl += h["net_qty"] * (price - prev[0])
            except sqlite3.Error:
                pass
            conn.close()
    overall_pnl = total_current - total_invested
    overall_pnl_pct = (overall_pnl / total_invested * 100) if total_invested else 0
    day_pnl_pct = (day_pnl / total_current * 100) if total_current else 0
    date = record_snapshot(
        total_invested,
        total_current,
        overall_pnl,
        overall_pnl_pct,
        day_pnl,
        day_pnl_pct,
    )
    print(f"{GREEN}Snapshot saved for {date}.{RESET}")
    print(f"  Invested: {fmt_inr(total_invested)}")
    print(f"  Value:    {fmt_inr(total_current)}")
    print(f"  P&L:      {color_pnl(overall_pnl)} ({color_pnl_pct(overall_pnl_pct)})")


def cmd_history(args):
    snapshots = get_snapshots(args.days)
    if not snapshots:
        print(f"{YELLOW}No snapshots found. Run 'snapshot' first.{RESET}")
        return
    table_data = []
    for s in snapshots:
        table_data.append(  # noqa: PG-APPEND
            [
                s["date"],
                fmt_inr(s["total_invested"]),
                fmt_inr(s["total_current"]),
                color_pnl(s["overall_pnl"]),
                color_pnl_pct(s["overall_pnl_pct"]),
                color_pnl(s["day_pnl"]),
            ]
        )
    headers = ["Date", "Invested", "Current", "P&L", "P&L%", "Day P&L"]
    if tabulate:
        print(tabulate(table_data, headers=headers, tablefmt="simple"))
    else:
        print("  ".join(headers))
        for row in table_data:
            print("  ".join(str(c) for c in row))
    if len(snapshots) >= 7:
        latest = snapshots[0]
        week_ago = snapshots[6]
        diff = latest["total_current"] - week_ago["total_current"]
        diff_str = color_pnl(diff)
        print(f"\nSince last week: {diff_str}")


def cmd_export(args):
    format_type = args.format
    if format_type == "csv":
        os.makedirs(EXPORTS_DIR, exist_ok=True)
        filename = (
            args.filename or f"exports/portfolio_{datetime.now():%Y%m%d_%H%M%S}.csv"
        )
        filepath = (
            filename
            if os.path.isabs(filename)
            else os.path.join(
                os.path.dirname(os.path.dirname(os.path.abspath(__file__))), filename
            )
        )
        os.makedirs(os.path.dirname(os.path.abspath(filepath)), exist_ok=True)
        rows = build_portfolio_rows(use_live=False, sort_col=None)
        if not rows:
            return
        with open(filepath, "w", newline="") as f:
            fieldnames = [
                "symbol",
                "qty",
                "avg",
                "ltp",
                "invested",
                "current",
                "pnl",
                "pnl_pct",
                "day_pnl",
                "day_pnl_pct",
                "del_pct",
                "del_trend",
                "vs_sma_pct",
                "vs_52w_high_pct",
                "vs_52w_low_pct",
                "pe",
                "sector",
                "market_cap",
            ]
            w = csv.DictWriter(f, fieldnames=fieldnames)
            w.writeheader()
            for r in rows:
                r_clean = {k: v for k, v in r.items() if k in fieldnames}
                w.writerow(r_clean)
        print(f"{GREEN}Exported {len(rows)} holdings to {filepath}{RESET}")


def cmd_performance(args):
    rows = build_portfolio_rows(use_live=False, sort_col=None)
    if not rows:
        return
    rows.sort(key=lambda r: r["pnl"])
    total_invested = sum(r["invested"] for r in rows)
    total_current = sum(r["current"] for r in rows)
    table_data = []
    for r in rows:
        weight = (r["current"] / total_current * 100) if total_current else 0
        table_data.append(  # noqa: PG-APPEND
            [
                r["symbol"],
                fmt_inr(r["invested"]),
                fmt_inr(r["current"]),
                color_pnl(r["pnl"]),
                color_pnl_pct(r["pnl_pct"]),
                f"{weight:.1f}%",
                color_pnl(r["day_pnl"]),
            ]
        )
    headers = ["Symbol", "Invested", "Current", "P&L", "P&L%", "Weight", "Day P&L"]
    if tabulate:
        print(f"\n{BOLD}Per-Stock Performance{RESET}")
        print(tabulate(table_data, headers=headers, tablefmt="simple"))
    else:
        print("  ".join(headers))
        for row in table_data:
            print("  ".join(str(c) for c in row))
    sec_alloc = get_sector_allocation(rows)
    if sec_alloc:
        sec_data = [
            [s["sector"], s["count"], fmt_inr(s["total_value"]), f"{s['weight_pct']}%"]
            for s in sec_alloc
        ]
        total_str = ["", "", fmt_inr(total_current), "100%"]
        if tabulate:
            print(f"\n{BOLD}Sector Allocation{RESET}")
            print(
                tabulate(
                    sec_data + [total_str],
                    headers=["Sector", "Holdings", "Value", "Weight%"],
                    tablefmt="simple",
                )
            )
        else:
            print(f"\n{'  '.join(['Sector','Holdings','Value','Weight%'])}")
            for row in sec_data + [total_str]:
                print("  ".join(str(c) for c in row))


def cmd_scanner(args):
    holdings = get_all_holdings()
    if not holdings:
        print(f"{YELLOW}No holdings in portfolio.{RESET}")
        return
    overlap = get_scanner_overlap(holdings)
    has_signal = any(
        any(v is not None for v in data.values()) for data in overlap.values()
    )
    if not has_signal:
        print(f"{YELLOW}No scanner signals detected for your holdings.{RESET}")
        return
    scanner_names = [
        "Trigger",
        "InvisHand",
        "FloatExh",
        "Wyckoff",
        "OpFinger",
        "LiqFlip",
        "Darvas",
        "Launchpad",
        "SeasDel",
    ]
    headers = ["Symbol"] + scanner_names
    table_data = []
    for symbol in sorted(overlap.keys()):
        data = overlap[symbol]
        row = [symbol]
        any_hit = False
        for sn in scanner_names:
            raw = data.get(sn)
            if raw is None:
                row.append(f"{YELLOW}{DASH}{RESET}")  # noqa: PG-APPEND
            else:
                any_hit = True
                if sn == "Trigger":
                    grade = raw.get("grade", raw.get("trigger_signal", ""))
                    row.append(  # noqa: PG-APPEND
                        f"{GREEN}{grade}{RESET}" if grade else f"{GREEN}{CHECK}{RESET}"
                    )
                elif sn == "InvisHand":
                    score = raw.get("ih_score")
                    if score:
                        row.append(f"{GREEN}{score:.0f}{RESET}")  # noqa: PG-APPEND
                    else:
                        row.append(f"{GREEN}{CHECK}{RESET}")  # noqa: PG-APPEND
                elif sn == "FloatExh":
                    util = raw.get("float_util_pct", 0)
                    if util >= 60:
                        row.append(f"{GREEN}high{RESET}")  # noqa: PG-APPEND
                    elif util >= 30:
                        row.append(f"{YELLOW}mod{RESET}")  # noqa: PG-APPEND
                    else:
                        row.append(f"{CYAN}low{RESET}")  # noqa: PG-APPEND
                elif sn == "Wyckoff":
                    phase = raw.get("scheme", raw.get("phase", ""))
                    row.append(  # noqa: PG-APPEND
                        f"{CYAN}{phase}{RESET}" if phase else f"{GREEN}{CHECK}{RESET}"
                    )
                elif sn == "OpFinger":
                    cr = raw.get("compression_ratio", 0)
                    if cr and cr > 2:
                        row.append(f"{GREEN}high{RESET}")  # noqa: PG-APPEND
                    elif cr and cr > 1:
                        row.append(f"{YELLOW}med{RESET}")  # noqa: PG-APPEND
                    else:
                        row.append(f"{CYAN}low{RESET}")  # noqa: PG-APPEND
                elif sn in ("LiqFlip", "Darvas", "Launchpad"):
                    row.append(f"{GREEN}{CHECK}{RESET}")  # noqa: PG-APPEND
                elif sn == "SeasDel":
                    month = raw.get("current_month", "")
                    row.append(  # noqa: PG-APPEND
                        f"{CYAN}{month}{RESET}" if month else f"{GREEN}{CHECK}{RESET}"
                    )
        if any_hit:
            table_data.append(row)  # noqa: PG-APPEND
    if not table_data:
        print(f"{YELLOW}No scanner signals detected for your holdings.{RESET}")
        return
    print(f"\n{BOLD}Scanner Overlap {DASH} Your Holdings{RESET}")
    if tabulate:
        print(tabulate(table_data, headers=headers, tablefmt="simple"))
    else:
        print("  ".join(headers))
        for row in table_data:
            print("  ".join(str(c) for c in row))


def cmd_alerts(args):
    holdings = get_all_holdings()
    if not holdings:
        print(f"{YELLOW}No holdings in portfolio.{RESET}")
        return
    alerts = get_delivery_alerts(holdings)
    if not alerts:
        print(
            f"{GREEN}No active alerts. All holdings have normal delivery patterns.{RESET}"
        )
        return
    sev_order = {"high": 0, "warning": 1, "info": 2}
    alerts.sort(key=lambda a: (sev_order.get(a["severity"], 99), a["symbol"]))
    print(f"\n{BOLD}Delivery Anomaly Alerts{RESET}")
    sev_colors = {"high": RED, "warning": YELLOW, "info": CYAN}
    for a in alerts:
        color = sev_colors.get(a["severity"], RESET)
        label = a["alert_type"].ljust(18)
        print(
            f"  {color}[{a['severity'].upper()}]{RESET} {styled(label, BOLD)} {a['symbol']}"
        )
        print(f"         {a['detail']}")


def cmd_risk(args):
    print(f"\n{BOLD}Portfolio Risk Report{RESET}")
    conc = get_concentration_risk()
    HR = "-" * 55

    def _hline():
        try:
            print(f"  {'-' * 55}")
        except UnicodeEncodeError:
            print(f"  {'-' * 55}")

    if conc:
        print(f"\n  Concentration Risk")
        print(f"  {HR}")
        top3_label = styled(
            f'{conc["top3_pct"]}%', YELLOW if conc["top3_pct"] >= 50 else GREEN
        )
        print(f"  Top 3 holdings: {top3_label} of portfolio")
        for h in conc["top3_holdings"]:
            print(f"    {h['symbol']}: {fmt_inr(h['value'])} ({h['pct']}%)")
        if conc["top3_pct"] >= 50:
            print(f"  {RED}!! Concentration is high. Consider diversifying.{RESET}")
        else:
            print(f"  {GREEN}Concentration is within reasonable range.{RESET}")
    draw = get_drawdown_metrics()
    if draw:
        print(f"\n  Drawdown Analysis (from snapshots history)")
        _hline()
        print(f"  Peak value:     {fmt_inr(draw['peak_value'])} ({draw['peak_date']})")
        print(f"  Current value:  {fmt_inr(draw['current_value'])}")
        print(
            f"  Max drawdown:   {color_pnl_pct(draw['drawdown_pct'])} ({fmt_inr(abs(draw['drawdown_amount']))})"
        )
        print(f"  Days from peak: {draw['days_from_peak']}")
    else:
        print(f"\n  Drawdown Analysis")
        _hline()
        print(
            f"  {YELLOW}Not enough history. Run 'snapshot' daily to build history.{RESET}"
        )
    mcap = get_allocation_by_mcap()
    if mcap:
        print(f"\n  Allocation by Market Cap")
        _hline()
        labels = {
            "large": "Large Cap",
            "mid": "Mid Cap",
            "small": "Small Cap",
            "unknown": "Unknown",
        }
        for k in ["large", "mid", "small", "unknown"]:
            v = mcap[k]
            if v["count"]:
                print(f"  {labels[k]:12s} {v['count']:2d} holdings  ({v['pct']:.1f}%)")
    vol = get_volatility_metrics()
    if vol:
        print(f"\n  Volatility (from daily snapshots)")
        _hline()
        print(f"  30-day volatility: {vol['daily_vol_pct']}% daily")
        print(
            f"  Max single-day gain: {color_pnl(vol['max_gain'])} ({vol['gain_date']})"
        )
        print(
            f"  Max single-day loss: {color_pnl(vol['max_loss'])} ({vol['loss_date']})"
        )
    else:
        print(f"\n  Volatility")
        _hline()
        print(
            f"  {YELLOW}Not enough history. Run 'snapshot' daily to build history.{RESET}"
        )
    div = get_diversification_score()
    if div:
        score_color = (
            GREEN if div["score"] >= 60 else (YELLOW if div["score"] >= 40 else RED)
        )
        print(
            f"\n  Diversification Score: {score_color}{div['score']}/100{RESET} ({div['rating']})"
        )
        print(f"  {div['details']}")


def cmd_status(args):
    holdings = get_all_holdings()
    if not holdings:
        print(f"{YELLOW}No holdings in portfolio. Use 'import' or 'add' first.{RESET}")
        return

    last_refresh = _get_portfolio_meta("last_refresh")
    holdings_count = len(holdings)

    fund_avail = 0
    try:
        conn = get_portfolio_conn()
        row = conn.execute("SELECT COUNT(*) FROM fundamental_cache").fetchone()
        fund_avail = row[0] if row else 0
        conn.close()
    except Exception:
        pass

    latest_eod_date = None
    try:
        conn = get_portfolio_conn()
        row = conn.execute(
            "SELECT latest_date FROM price_cache ORDER BY latest_date DESC LIMIT 1"
        ).fetchone()
        if row:
            latest_eod_date = row[0]
        conn.close()
    except Exception:
        pass

    fund_pct = round(fund_avail / holdings_count * 100) if holdings_count else 0

    print(f"\n{BOLD}Portfolio Status{RESET}")
    print(f"{DASH * 53}")
    print(f"Holdings:     {holdings_count} symbols")
    print(f"Last refresh: {last_refresh or 'Never'} (auto)")
    print(f"Prices from:  {latest_eod_date or 'N/A'} (EOD)")
    print(f"Fundamentals: {fund_avail} of {holdings_count} available ({fund_pct}%)")
    print(f"{DASH * 53}")
    print(f"Next auto-refresh: After next bhavcopy ingestion (~18:30 IST tomorrow)")


def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="MYRA Portfolio Tracker - CLI tool for managing stock holdings.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python tools/portfolio.py import "Holdings_14-Jun-2026_10.47.43.xlsx"
  python tools/portfolio.py view --compact
  python tools/portfolio.py view --detailed
  python tools/portfolio.py view --live
  python tools/portfolio.py add RELIANCE 10 2500
  python tools/portfolio.py sell RELIANCE 5 2600
  python tools/portfolio.py snapshot
  python tools/portfolio.py history --days 60
  python tools/portfolio.py export csv
  python tools/portfolio.py performance
  python tools/portfolio.py scanner
  python tools/portfolio.py alerts
  python tools/portfolio.py risk
        """,
    )
    sub = parser.add_subparsers(dest="command")
    sub.required = True

    p_import = sub.add_parser("import", help="Import holdings from broker XLSX")
    p_import.add_argument("path", help="Path to XLSX file")

    p_view = sub.add_parser("view", help="View portfolio with prices and P&L")
    p_view.add_argument(
        "--live", action="store_true", help="Fetch live prices via yfinance"
    )
    p_view.add_argument(
        "--sort",
        choices=["symbol", "value", "pnl", "day_pnl", "pnl_pct"],
        default=None,
        help="Sort column",
    )
    p_view.add_argument(
        "--compact",
        action="store_true",
        help="Compact view: symbol, LTP, P&L%, alerts only",
    )
    p_view.add_argument(
        "--detailed", action="store_true", help="Show all enriched columns"
    )

    p_add = sub.add_parser("add", help="Add a holding")
    p_add.add_argument("symbol")
    p_add.add_argument("qty", type=int)
    p_add.add_argument("price", type=float)
    p_add.add_argument("--category", default="NSE EQ")

    p_sell = sub.add_parser("sell", help="Sell shares")
    p_sell.add_argument("symbol")
    p_sell.add_argument("qty", type=int)
    p_sell.add_argument("price", type=float)

    p_update = sub.add_parser("update", help="Update holding fields")
    p_update.add_argument("symbol")
    p_update.add_argument("--ltp", type=float, help="Set LTP as new avg price")
    p_update.add_argument("--qty", type=int, help="Update quantity")
    p_update.add_argument("--avg", type=float, help="Update avg price")

    p_snap = sub.add_parser("snapshot", help="Record daily portfolio snapshot")
    p_snap.add_argument("--date", help="Override date (YYYY-MM-DD)")

    p_hist = sub.add_parser("history", help="View historical snapshots")
    p_hist.add_argument("--days", type=int, default=30, help="Number of days")

    p_export = sub.add_parser("export", help="Export portfolio")
    p_export.add_argument("format", choices=["csv"], default="csv", nargs="?")
    p_export.add_argument("--filename", "-f", help="Output filename")

    p_perf = sub.add_parser("performance", help="Per-stock and sector breakdown")

    p_scanner = sub.add_parser(
        "scanner", help="Cross-reference holdings with MYRA scanner results"
    )
    p_alerts = sub.add_parser(
        "alerts", help="Show delivery anomaly alerts for your holdings"
    )

    p_risk = sub.add_parser(
        "risk", help="Portfolio risk metrics (concentration, drawdown, volatility)"
    )

    p_status = sub.add_parser(
        "status", help="Show portfolio refresh state and metadata"
    )

    args = parser.parse_args()

    handlers = {
        "import": cmd_import,
        "view": cmd_view,
        "add": cmd_add,
        "sell": cmd_sell,
        "update": cmd_update,
        "snapshot": cmd_snapshot,
        "history": cmd_history,
        "export": cmd_export,
        "performance": cmd_performance,
        "scanner": cmd_scanner,
        "alerts": cmd_alerts,
        "risk": cmd_risk,
        "status": cmd_status,
    }

    handler = handlers.get(args.command)
    if handler:
        handler(args)


if __name__ == "__main__":
    main()
